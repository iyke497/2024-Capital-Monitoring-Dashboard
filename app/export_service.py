"""
Export service for survey responses
Handles exporting survey data to Excel format
"""
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from flask import current_app
import json


class ExportService:
    """Service for exporting survey responses to Excel"""
    
    # Column headers mapping - order matters for Excel output
    EXPORT_COLUMNS = [
        ("PARENT MINISTRY", "parent_ministry"),
        ("NAME OF MDA", "mda_name"),
        ("PROJECT NAME", "project_name"),
        ("SUB-PROJECT/ACTIVITY", "sub_projects"),
        ("STRATEGIC OBJECTIVES IN ACCORDANCE WITH NDP", "strategic_objective"),
        ("KEY PERFORMANCE INDICATORS", "key_performance_indicators"),
        ("PROJECT TYPE", "project_type"),
        ("PROJECT DELIVERABLES", "project_deliverables"),
        ("PROJECT EXECUTION", "execution_method"),
        ("CONTRACTOR RC NUMBERS", "contractor_rc_numbers"),
        ("CONTRACTOR NAME", "contractor_name"),
        ("CERTIFICATE OF AWARD", "award_certificate"),
        ("PROJECT CATEGORIZATION", "project_categorisation"),
        ("PROJECT APPROPRIATION", "project_appropriation_2024"),
        ("AMOUNT RELEASED", "amount_released_2024"),
        ("AMOUNT UTILIZED", "amount_utilized_2024"),
        ("TOTAL COST OF PROJECT PLANNED", "total_cost_planned"),
        ("TOTAL FINANCIAL COMMITMENT SINCE INCEPTION", "total_financial_commitment"),
        ("PROJECT PICTURES", "project_pictures"),
        ("OTHER RELEVANT DOCUMENTS", "other_documents"),
        ("JOB COMPLETION CERTIFICATE ISSUED", "completion_cert_issued"),
        ("JOB COMPLETION CERTIFICATE", "job_completion_certificate"),
        ("TOTAL AMOUNT IN APPROVED PROJECT COMPLETION CERTIFICATE", "completion_cert_amount"),
        ("PROJECT STATUS", "project_status"),
        ("START DATE", "start_date"),
        ("END DATE", "end_date"),
        ("PERCENTAGE COMPLETED", "percentage_completed"),
        ("LIST PROJECT ACHIEVEMENTS", "project_achievements"),
        ("GEOLOCATIONS", "geolocations"),
        ("STATE", "state"),
        ("LGA", "lga"),
        ("WARD", "ward"),
        ("WHAT ARE THE CHALLENGES AND RECOMMENDATIONS", "challenges_recommendations"),
    ]
    
    # Fields that contain file URLs
    FILE_FIELDS = ['award_certificate', 'project_pictures', 'other_documents', 'job_completion_certificate']
    
    @classmethod
    def format_cell_value(cls, value, field_name):
        """Format cell values based on field type"""
        if value is None:
            return ""
        
        # Handle boolean fields
        if field_name == 'completion_cert_issued':
            return "Yes" if value else "No"
        
        # Handle JSON/file fields - stored as arrays of URL strings
        if field_name in cls.FILE_FIELDS:
            # If it's already a list (JSON deserialized)
            if isinstance(value, list):
                # Filter out empty/None values and join with newlines
                urls = [str(url).strip() for url in value if url]
                result = '\n'.join(urls) if urls else ""
                current_app.logger.info(f"FILE FIELD {field_name}: Got list with {len(urls)} URLs")
                return result
            
            # If it's a string, try to parse it as JSON
            elif isinstance(value, str):
                try:
                    parsed = json.loads(value)
                    if isinstance(parsed, list):
                        urls = [str(url).strip() for url in parsed if url]
                        current_app.logger.info(f"FILE FIELD {field_name}: Parsed string to list with {len(urls)} URLs")
                        return '\n'.join(urls) if urls else ""
                except (json.JSONDecodeError, TypeError):
                    # Not valid JSON, return as-is (might be a single URL)
                    current_app.logger.info(f"FILE FIELD {field_name}: Returning string as-is")
                    return value.strip()
            
            # Fallback for other types
            current_app.logger.warning(f"FILE FIELD {field_name}: Unexpected type {type(value)}, value: {value}")
            return str(value) if value else ""
        
        # Handle numeric fields
        if field_name in ['project_appropriation_2024', 'amount_released_2024', 
                         'amount_utilized_2024', 'total_cost_planned', 
                         'total_financial_commitment', 'completion_cert_amount']:
            try:
                return float(value)
            except (ValueError, TypeError):
                return ""
        
        # Handle date fields
        if field_name in ['start_date', 'end_date']:
            if hasattr(value, 'isoformat'):
                return value.isoformat()
            return str(value)
        
        # Default string conversion
        return str(value) if value else ""
    
    @classmethod
    def export_to_excel(cls, responses, filename=None):
        """
        Export survey responses to Excel file
        
        Args:
            responses: List of SurveyResponse model instances
            filename: Optional filename (default: auto-generated with timestamp)
            
        Returns:
            BytesIO: Excel file as bytes buffer
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'survey_responses_{timestamp}.xlsx'
        
        current_app.logger.info(f"Starting export of {len(responses)} responses")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Survey Responses"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_idx, (header, _) in enumerate(cls.EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style
        
        # Write data rows
        for row_idx, response in enumerate(responses, start=2):
            for col_idx, (_, field_name) in enumerate(cls.EXPORT_COLUMNS, start=1):
                # Get value from model
                value = getattr(response, field_name, None)
                
                # Format and write cell
                formatted_value = cls.format_cell_value(value, field_name)
                cell = ws.cell(row=row_idx, column=col_idx, value=formatted_value)
                cell.border = border_style
                
                # Apply number format for financial fields
                if field_name in ['project_appropriation_2024', 'amount_released_2024', 
                                 'amount_utilized_2024', 'total_cost_planned', 
                                 'total_financial_commitment', 'completion_cert_amount']:
                    cell.number_format = '#,##0.00'
                
                # Enable text wrapping for file/document fields with multiple URLs
                if field_name in cls.FILE_FIELDS:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Auto-adjust column widths
        for col_idx, (header, field_name) in enumerate(cls.EXPORT_COLUMNS, start=1):
            column_letter = get_column_letter(col_idx)
            
            # Calculate max width based on header and sample data
            max_length = len(header)
            for row_idx in range(2, min(ws.max_row + 1, 102)):  # Check first 100 rows
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    # For multiline cells, use the longest line
                    if isinstance(cell_value, str) and '\n' in cell_value:
                        lines = cell_value.split('\n')
                        max_length = max(max_length, max(len(line) for line in lines))
                    else:
                        max_length = max(max_length, len(str(cell_value)))
            
            # Set width (with limits) - give more space to URL fields
            if field_name in cls.FILE_FIELDS:
                adjusted_width = min(max_length + 2, 80)  # Wider for URLs
            else:
                adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze the header row
        ws.freeze_panes = "A2"
        
        current_app.logger.info("Excel file generated successfully")
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output, filename
    
    @classmethod
    def export_filtered_responses(cls, query_filters=None):
        """
        Export responses with optional filters
        
        Args:
            query_filters: Dictionary of filter criteria
            
        Returns:
            BytesIO: Excel file as bytes buffer
        """
        from app.models import SurveyResponse
        
        query = SurveyResponse.query
        
        # Apply filters if provided
        if query_filters:
            if 'survey_type' in query_filters:
                query = query.filter_by(survey_type=query_filters['survey_type'])
            
            if 'parent_ministry' in query_filters:
                query = query.filter_by(parent_ministry=query_filters['parent_ministry'])
            
            if 'state' in query_filters:
                query = query.filter_by(state=query_filters['state'])
            
            if 'project_status' in query_filters:
                query = query.filter_by(project_status=query_filters['project_status'])
            
            if 'start_date' in query_filters:
                query = query.filter(SurveyResponse.created >= query_filters['start_date'])
            
            if 'end_date' in query_filters:
                query = query.filter(SurveyResponse.created <= query_filters['end_date'])
        
        # Order by parent ministry, then MDA for better organization
        responses = query.order_by(
            SurveyResponse.parent_ministry,
            SurveyResponse.mda_name,
            SurveyResponse.project_name
        ).all()
        
        return cls.export_to_excel(responses)